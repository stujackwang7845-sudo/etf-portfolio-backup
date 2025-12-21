"""
GitHub Actions 執行腳本 - 擷取並儲存 ETF 投資組合資料
在 GitHub 雲端伺服器上執行，無需本地電腦開機
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from datetime import datetime
import time
import re
import openpyxl
from openpyxl.styles import Font, Alignment


def fetch_etf_data():
    """擷取 ETF 投資組合資料"""
    
    # 設定 Chrome 選項
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    
    # 使用系統的 chromium-chromedriver
    service = Service('/usr/bin/chromedriver')
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    try:
        url = "https://www.ezmoney.com.tw/ETF/Fund/Info?fundCode=49YTW"
        print(f"正在載入頁面: {url}")
        driver.get(url)
        time.sleep(3)
        
        # 點擊「基金投資組合」分頁
        try:
            portfolio_tab = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.LINK_TEXT, "基金投資組合"))
            )
            portfolio_tab.click()
            print("已點擊「基金投資組合」分頁")
            time.sleep(2)
        except:
            print("無需點擊分頁")
        
        # 擷取資料日期 - 從頁面文字中尋找「資料日期」標籤
        page_text = driver.find_element(By.TAG_NAME, 'body').text
        
        # 尋找「資料日期:YYYY/MM/DD」格式
        date_pattern = r'資料日期[：:]\s*(\d{3,4})[/-](\d{1,2})[/-](\d{1,2})'
        match = re.search(date_pattern, page_text)
        
        if match:
            year_or_roc = match.group(1)
            month = match.group(2)
            day = match.group(3)
            
            # 如果是民國年（3位數），轉換為西元年
            if len(year_or_roc) == 3:
                year = str(int(year_or_roc) + 1911)
            else:
                year = year_or_roc
            
            data_date = f"{year}/{month.zfill(2)}/{day.zfill(2)}"
            print(f"✅ 找到資料日期: {data_date}")
        else:
            # 備用：使用當前日期
            data_date = datetime.now().strftime('%Y/%m/%d')
            print(f"⚠️  無法找到資料日期，使用當前日期: {data_date}")
        
        # 擷取持股資料
        page_text = driver.find_element(By.TAG_NAME, 'body').text
        lines = [line.strip() for line in page_text.split('\n') if line.strip()]
        
        holdings = []
        stock_pattern = r'^(\d{4})\s+(.+?)\s+([\d,]+)\s+([\d.]+)%'
        
        # 找到表頭
        for i, line in enumerate(lines):
            if '股票代號' in line and '股票名稱' in line and '股數' in line:
                start_idx = i + 1
                print(f"找到股票表格標題於第 {i} 行")
                
                # 解析持股
                for j in range(start_idx, len(lines)):
                    match = re.match(stock_pattern, lines[j])
                    if match:
                        stock_code = match.group(1)
                        stock_name = match.group(2).strip()
                        shares_text = match.group(3)
                        weight_text = match.group(4)
                        
                        shares = float(shares_text.replace(',', ''))
                        weight = float(weight_text)
                        
                        holdings.append({
                            'stock_code': stock_code,
                            'stock_name': stock_name,
                            'shares': shares,
                            'weight': weight
                        })
                    elif '友善列印' in lines[j] or '匯出' in lines[j]:
                        break
                
                break
        
        print(f"成功擷取 {len(holdings)} 筆持股資料")
        
        # 擷取基金資產資訊
        fund_info = {}
        try:
            # 尋找基金資產相關數據
            lines_text = '\n'.join(lines)
            
            # 為了調試，打印一些關鍵文字
            print("\n=== 調試：尋找資產配置資訊 ===")
            
            # === 基金資產 ===
            # 淨資產 - 更寬鬆的匹配
            for pattern in [
                r'淨資產\s*[:：]?\s*(NTD\s*-?[\d,]+)',
                r'淨資產[^\d]+(NTD\s*-?[\d,]+)',
            ]:
                match = re.search(pattern, lines_text, re.IGNORECASE)
                if match:
                    fund_info['net_asset'] = match.group(1)
                    print(f"✓ 淨資產: {match.group(1)}")
                    break
            
            # 流通在外單位數
            for pattern in [
                r'流通在外單位數\s*[:：]?\s*([\d,]+)',
                r'流通在外單位數[^\d]+([\d,]+)',
            ]:
                match = re.search(pattern, lines_text)
                if match:
                    fund_info['outstanding_units'] = match.group(1)
                    print(f"✓ 流通在外單位數: {match.group(1)}")
                    break
            
            # 每單位淨值
            for pattern in [
                r'每單位淨值\s*[:：]?\s*(NTD\s*-?[\d.]+)',
                r'每單位淨值[^\d]+(NTD\s*-?[\d.]+)',
            ]:
                match = re.search(pattern, lines_text, re.IGNORECASE)
                if match:
                    fund_info['nav'] = match.group(1)
                    print(f"✓ 每單位淨值: {match.group(1)}")
                    break
            
            # === 尋找表格形式的資產配置 ===
            # 尋找「項目」「金額」「權重」表格
            items_to_find = {
                '現金': ('cash', 'cash_weight'),
                '期貨保證金': ('futures_margin', 'futures_margin_weight'),
                '申贖應付款': ('subscription_payable', 'subscription_payable_weight'),
                '應收付證券款': ('securities_payable', 'securities_payable_weight'),
                '期貨': ('futures', 'futures_weight'),
                '股票': ('stocks_value', 'stocks_weight'),
            }
            
            for item_name, (value_key, weight_key) in items_to_find.items():
                # 嘗試找到「項目名稱 金額 權重」的模式
                # 注意：百分比可能是 "0.56%" 或 "0.56 %"（有空格），且可能為負數
                patterns = [
                    rf'{item_name}\s+(NTD\s*-?[\d,]+)\s+(-?[\d.]+)\s*%',  # 允許 % 前有空格，允許負數
                    rf'{item_name}[^\n]*?(NTD\s*-?[\d,]+)[^\d]+(-?[\d.]+)\s*%',
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, lines_text, re.IGNORECASE)
                    if match:
                        fund_info[value_key] = match.group(1)
                        fund_info[weight_key] = match.group(2) + '%'
                        print(f"✓ {item_name}: {match.group(1)}, {match.group(2)}%")
                        break
            
            print("=== 調試結束 ===\n")
            
            extracted_count = len([k for k in fund_info.keys()])
            print(f"✅ 成功擷取 {extracted_count} 項基金資產資訊")
        except Exception as e:
            print(f"⚠️  基金資產資訊擷取失敗: {e}")
            import traceback
            traceback.print_exc()
        
        return {
            'date': data_date,
            'holdings': holdings,
            'fund_info': fund_info
        }
        
    finally:
        driver.quit()


def save_to_excel(portfolio_data):
    """儲存為 Excel 格式"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 資料日期（使用西元年）
    date_str = portfolio_data['date']
    date_obj = datetime.strptime(date_str, '%Y/%m/%d')
    formatted_date = f"{date_obj.year}/{date_obj.month:02d}/{date_obj.day:02d}"
    
    # 建立 Excel 結構
    ws['A1'] = f'資料日期：{formatted_date}'
    ws['A1'].font = Font(bold=True)
    # 基金資產資訊（從 portfolio_data 取得，如果沒有則使用預設值）
    fund_info = portfolio_data.get('fund_info', {})
    
    ws['A3'] = '基金資產'
    ws['A3'].font = Font(bold=True)
    
    ws['A4'] = '淨資產'
    ws['B4'] = fund_info.get('net_asset', 'NTD 42,575,942,188')
    
    ws['A5'] = '流通在外單位數'
    ws['B5'] = fund_info.get('outstanding_units', '2,596,709,000')
    
    ws['A6'] = '每單位淨值'
    ws['B6'] = fund_info.get('nav', 'NTD 16.40')
    
    # 資產配置
    ws['A8'] = '項目'
    ws['B8'] = '金額'
    ws['C8'] = '權重'
    
    ws['A9'] = '期貨(名目本金)'
    ws['B9'] = fund_info.get('futures', 'NTD 0')
    ws['C9'] = fund_info.get('futures_weight', '0%')
    
    total_weight = sum(h['weight'] for h in portfolio_data['holdings'])
    ws['A10'] = '股票'
    ws['B10'] = fund_info.get('stocks_value', 'NTD 40,529,643,608')
    ws['C10'] = f'{total_weight:.2f}%'
    
    ws['A12'] = '項目'
    ws['B12'] = '金額'
    ws['C12'] = '權重'
    
    ws['A13'] = '現金'
    ws['B13'] = fund_info.get('cash', 'NTD 0')
    ws['C13'] = fund_info.get('cash_weight', '0%')
    
    ws['A14'] = '期貨保證金'
    ws['B14'] = fund_info.get('futures_margin', 'NTD 0')
    ws['C14'] = fund_info.get('futures_margin_weight', '0%')
    
    ws['A15'] = '申贖應付款'
    ws['B15'] = fund_info.get('subscription_payable', 'NTD 0')
    ws['C15'] = fund_info.get('subscription_payable_weight', '0%')
    
    ws['A16'] = '應收付證券款'
    ws['B16'] = fund_info.get('securities_payable', 'NTD 0')
    ws['C16'] = fund_info.get('securities_payable_weight', '0%')
    
    # 股票資料
    ws['A19'] = '股票'
    ws['A19'].font = Font(bold=True)
    
    ws['A20'] = '股票代號'
    ws['B20'] = '股票名稱'
    ws['C20'] = '股數'
    ws['D20'] = '持股權重'
    
    for cell in ws[20]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 填入持股
    for idx, holding in enumerate(portfolio_data['holdings'], start=21):
        ws[f'A{idx}'] = holding['stock_code']
        ws[f'B{idx}'] = holding['stock_name']
        ws[f'C{idx}'] = f"{holding['shares']:,.0f}"
        ws[f'D{idx}'] = f"{holding['weight']:.2f}%"
    
    # 調整欄寬
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    
    # 儲存檔案
    date_str = portfolio_data['date'].replace('/', '')
    filename = f"ETF_Investment_Portfolio_{date_str}.xlsx"
    wb.save(filename)
    print(f"Excel 已儲存: {filename}")
    
    return filename


if __name__ == '__main__':
    print("="*60)
    print("GitHub Actions - ETF 投資組合自動擷取")
    print("="*60)
    
    try:
        # 擷取資料
        print("\n[1/2] 擷取 ETF 投資組合資料...")
        data = fetch_etf_data()
        
        # 儲存 Excel
        print("\n[2/2] 儲存為 Excel 格式...")
        filename = save_to_excel(data)
        
        print("\n" + "="*60)
        print("✅ 執行成功！")
        print(f"資料日期: {data['date']}")
        print(f"持股數量: {len(data['holdings'])}")
        print(f"檔案名稱: {filename}")
        print("="*60)
        
    except Exception as e:
        print(f"\n❌ 執行失敗: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
